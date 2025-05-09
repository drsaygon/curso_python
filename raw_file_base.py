from abc import ABCMeta
from csv import DictWriter, writer
from io import BytesIO, StringIO
from openpyxl import load_workbook
import time
from concurrent.futures import ThreadPoolExecutor
from stratboxcore.library.etl.raw_data.constants import (
    CSV_FIELDNAMES,
    SCENARIO_ALL,
    SCENARIO_BASE,
    SCENARIO_ALTERNATIVE,
    HeaderConfig,
    HeaderType,
)
from stratboxcore.library.etl.raw_data.file_configs import FileReadConfigurations
from stratboxcore.library.etl.raw_data.section_handlers.section_base import (
    SectionBase,
    SectionValidationException,
)
from stratboxcore.library.etl.raw_data.file_validation import FileValidationException


class RawFileBase(metaclass=ABCMeta):
    """Base class for all files that can be uploaded."""
    def __init__(
        self,
        filename,
        file_type,
        scenario,
        exercise,
        section_handler,
        lookup,
        file_configs=None,
    ):
        if file_configs is None:
            file_configs = FileReadConfigurations()
        self._lookup = lookup
        self._file_configs = file_configs
        self.filename = filename
        self.file_type = file_type
        self.scenario = scenario
        self.exercise = exercise
        self.section_handler = section_handler
        self.template_config = None
        self.max_column = None
        self.row_configs = []
        self.warnings = []
        self.sections = []
        self._configure()

    def _configure(self):
        try:
            self.template_config = self._lookup[self.file_type]
            self.row_configs = self._file_configs[self.file_type]
        except KeyError:
            raise FileValidationException(
                "File type not recognised: {}".format(self.file_type), file=self
            )

        permitted_scenarios = []
        if SCENARIO_ALL in self.template_config.scenarios:
            permitted_scenarios.append(SCENARIO_ALL)
        if SCENARIO_BASE in self.template_config.scenarios:
            permitted_scenarios.append(self.exercise["base_scenario"])
        if SCENARIO_ALTERNATIVE in self.template_config.scenarios:
            alternative_scenarios = []
            for scenario in self.exercise["alternative_scenarios"]:
                if isinstance(scenario, dict):
                    if scenario.get("isActive", True) is True:
                        alternative_scenarios.append(scenario["name"])
                else:
                    alternative_scenarios.append(scenario)
            permitted_scenarios.extend(alternative_scenarios)
        if self.scenario not in permitted_scenarios:
            raise FileValidationException(
                "Scenario not in this file's permitted scenarios: {}".format(
                    ", ".join(permitted_scenarios)
                ),
                file=self,
            )

        if self.template_config.scenario_loop is not None:
            loop_start = self.template_config.scenario_loop[0]
            loop_stride = self.template_config.scenario_loop[1]
            loop_scenarios = self.exercise["alternative_scenarios"]
            if self.template_config.scenario_loop[2] == SCENARIO_ALL:
                loop_scenarios.insert(0, self.exercise["base_scenario"])
            for config_line in self.row_configs:
                if config_line.row_number < loop_start:
                    self.sections.append(
                        self.section_handler(
                            scenario=self.scenario,
                            exercise=self.exercise,
                            config=config_line,
                            template_config=self.template_config,
                        )
                    )
                else:
                    for index, scenario in enumerate(loop_scenarios):
                        self.sections.append(
                            self.section_handler(
                                scenario=(
                                    scenario["name"]
                                    if isinstance(scenario, dict)
                                    else scenario
                                ),
                                exercise=self.exercise,
                                config=config_line._replace(
                                    row_number=config_line.row_number
                                    + loop_stride * index
                                ),
                                template_config=self.template_config,
                            )
                        )
        else:
            self.sections = [
                self.section_handler(
                    scenario=self.scenario,
                    exercise=self.exercise,
                    config=config_line,
                    template_config=self.template_config,
                )
                for config_line in self.row_configs
            ]

    def _expected_header_values(self, header_config: HeaderConfig):
        if header_config.header_type == HeaderType.BUSINESS_UNIT:
            business_units = []
            for key in header_config.options:
                if key == "regulatory_capital_perimeters":
                    business_units.extend(self.exercise[key])
                else:
                    business_units.extend(self.exercise["business_units"][key])
            return [bu["name"] if isinstance(bu, dict) else bu for bu in business_units]
        elif header_config.header_type == HeaderType.PERIOD:
            periods = self.exercise["periods"]
            if header_config.options:
                periods = header_config.options(periods)
            return periods
        elif header_config.header_type == HeaderType.SCENARIO:
            return [self.exercise["base_scenario"]] + [
                sc["name"] if isinstance(sc, dict) else sc
                for sc in self.exercise["alternative_scenarios"]
            ]
        elif header_config.header_type == HeaderType.SPACER:
            return [None] * header_config.options
        else:
            raise Exception(
                "Unrecognised header_type {}".format(header_config.header_type)
            )

    def _validate_contents(self, sheet_title, rows):
        issues = []

        if self.template_config.title is not None and sheet_title != self.template_config.title:
            issues.append(
                'Sheet title mismatch: expected "{}", found "{}"'.format(
                    self.template_config.title, sheet_title
                )
            )

        headers = self.template_config.headers
        columns_between_groups = self.template_config.columns_between_groups

        if len(headers) == 0:
            pass
        elif len(headers) == 1:
            header_config = headers[0]
            expected_values = self._expected_header_values(header_config)
            self.max_column = (
                self.template_config.first_data_column
                + (len(expected_values) - 1) * columns_between_groups
            )
            try:
                values_row = rows[header_config.header_row - 1][:self.max_column]
            except IndexError:
                values_row = []

            for index, expected in enumerate(expected_values):
                column = self.template_config.first_data_column + index * columns_between_groups
                try:
                    found = values_row[column - 1]
                except IndexError:
                    found = None
                if str(found) != str(expected):
                    issues.append(
                        "Header mismatch on row {}, column {}: expected {}, found {}".format(
                            header_config.header_row, column, expected, found
                        )
                    )
        elif len(headers) == 2:
            first_expected = self._expected_header_values(headers[0])
            second_expected = self._expected_header_values(headers[1])

            self.max_column = (
                self.template_config.first_data_column
                + len(first_expected) * columns_between_groups
                - 1
            )

            try:
                first_row = rows[headers[0].header_row - 1][:self.max_column]
            except IndexError:
                first_row = []

            try:
                second_row = rows[headers[1].header_row - 1][:self.max_column]
            except IndexError:
                second_row = []

            for idx, expected in enumerate(first_expected):
                col = self.template_config.first_data_column + idx * columns_between_groups
                try:
                    found = first_row[col - 1]
                except IndexError:
                    found = None
                if str(found) != str(expected):
                    issues.append(
                        "Header mismatch on row {}, column {}: expected {}, found {}".format(
                            headers[0].header_row, col, expected, found
                        )
                    )

            if headers[1].header_type != HeaderType.SPACER:
                for first_idx, first_val in enumerate(first_expected):
                    base_col = self.template_config.first_data_column + first_idx * columns_between_groups
                    for second_idx, second_val in enumerate(second_expected):
                        col = base_col + second_idx
                        try:
                            found = second_row[col - 1]
                        except IndexError:
                            found = None
                        if str(found) != str(second_val):
                            issues.append(
                                "Header mismatch on row {}, column {}: expected {}, found {}".format(
                                    headers[1].header_row, col, second_val, found
                                )
                            )
        else:
            raise Exception(
                "Unexpected number of header configurations found ({})".format(len(headers))
            )
        return issues

    def _format_warnings(self):
        result = "Warning - {} blank cell(s) found in {} file for scenario {}".format(
            len(self.warnings), self.file_type, self.scenario
        )
        result += "".join(
            [
                "\n\t(Section starting row {}) Primary methodology {}, subsegment {}, metric {}, business unit {}, period {}".format(*warning)
                for warning in self.warnings
            ]
        )
        return result

    def _populate(self, sheet_title, rows, integration_test_ignore_empty_config=False):
        file_level_issues = self._validate_contents(sheet_title, rows)
        if file_level_issues:
            raise FileValidationException(
                "File is not correct format for uploading",
                file=self,
                issues=file_level_issues,
            )

        section_exceptions = []
        for section in self.sections:
            try:
                self.warnings.extend(
                    self._process_section(section, rows, integration_test_ignore_empty_config)
                )
            except SectionValidationException as ex:
                section_exceptions.append(ex)

        if section_exceptions:
            raise FileValidationException(
                "One or more sections had validation errors",
                file=self,
                issues=section_exceptions,
            )

        return self._format_warnings() if self.warnings else None

    def _process_section(self, section, rows, integration_test_ignore_empty_config):
        if (
            integration_test_ignore_empty_config
            and section.config.primary_methodology
            not in section.exercise["segmentsByMethodology"]
        ):
            return []

        try:
            return section.populate(rows, self.max_column)
        except SectionValidationException as ex:
            raise ex

    def populate(self, file):
        try:
            start_time = time.time()
            excel = load_workbook(file, data_only=True, read_only=True)
            end_time = time.time()

            sheet = excel.active
            rows = list(sheet.iter_rows(values_only=True))
            result = self._populate(sheet.title, rows)

            end_time2 = time.time()
            print(f'{round((end_time - start_time) + (end_time2 - start_time), 2)}')
            return result
        except Exception:
            raise FileValidationException("File is not an Excel file", file=self)

    def to_csv_string(self):
        buffer = StringIO()
        writer = DictWriter(buffer, fieldnames=CSV_FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        for section in self.sections:
            writer.writerows(section.extract_data())
        return buffer.getvalue()